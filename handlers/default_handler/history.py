from aiogram import types
from aiogram.utils.keyboard import InlineKeyboardBuilder
from loader import dp, bot
from aiogram import F
from aiogram.types import FSInputFile
from database.requests import *
from openpyxl import Workbook
import os.path


@dp.callback_query(F.data == "История")
async def search_response_history(callback: types.CallbackQuery):
    builder = InlineKeyboardBuilder()
    builder.row(types.InlineKeyboardButton(
        text="Назад", callback_data="/start")
    )
    builder.row(types.InlineKeyboardButton(
        text="Продолжить", callback_data="История запросов")
    )
    await callback.message.answer('Нажмите "Назад" для возврата в основное меню или \n'
                                  '"Продолжить" для получения истории запросов', reply_markup=builder.as_markup())


@dp.callback_query(F.data == "История запросов")
async def get_data_history(callback: types.CallbackQuery):
    builder = InlineKeyboardBuilder()
    builder.row(types.InlineKeyboardButton(
        text="Назад", callback_data="История")
    )
    await callback.message.answer("История запросов:", reply_markup=builder.as_markup())
    try:
        result = search_history()

        wb = Workbook()

        sheet = wb.active

        sheet.cell(row=1, column=1).value = "Город"
        sheet.cell(row=1, column=2).value = "Вакансия"
        sheet.cell(row=1, column=3).value = "дата"

        last_row = sheet.max_row + 1  # Определение номера строки для начала заполнения данных
        for row_index, row in enumerate(result, start=last_row):
            city, vacancy, date_request = row
            sheet.cell(row=row_index, column=1).value = city
            sheet.cell(row=row_index, column=2).value = vacancy
            sheet.cell(row=row_index, column=3).value = date_request

        for column in sheet.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    await callback.message.answer(f"Ошибка: {e}")
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        wb.save("history.xlsx")
        absolute_path = os.path.abspath("history.xlsx")
        document = FSInputFile(absolute_path)
        await bot.send_document(chat_id=callback.from_user.id, document=document)
    except Exception as e:
        await callback.message.answer(f"Ошибка: {e}")
